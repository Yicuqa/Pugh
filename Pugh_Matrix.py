import json
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from fpdf import FPDF
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


class PDF(FPDF):
    def __init__(self, title=None):
        super().__init__()
        self.title = title

    def header(self):
        if self.title:
            self.set_font('Arial', 'B', 12)
            self.cell(0, 10, self.title, 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

class CycleCheckbutton(tk.Label):
    def __init__(self, master=None, state_change_callback=None, **kw):
        super().__init__(master, **kw)
        self.configure(font=('Helvetica', 10), wraplength=100, bg='#f0f0f0', fg='black', relief='flat', pady=1, padx=1)
        self.values = [" ", "+", "-", "S"]
        self.current_value = 0
        self.configure(text=self.values[self.current_value], borderwidth=1, relief="groove")
        self.bind("<Button-1>", self.cycle_value)
        self.state_change_callback = state_change_callback


    def cycle_value(self, event=None):
        self.current_value = (self.current_value + 1) % len(self.values)
        self.configure(text=self.values[self.current_value])
        # Invoke the callback if it's set
        if self.state_change_callback:
            self.state_change_callback(self.get_state())

    def get_state(self):
        return self.values[self.current_value]

    def set_state(self, state):
        if state in self.values:
            self.current_value = self.values.index(state)
            self.configure(text=state)

    def get_score(self):
        value_scores = {"+": 1, "-": -1, "S": 0, " ": 0}
        return value_scores[self.values[self.current_value]]

class ToggleButton(tk.Frame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.config(borderwidth=2, relief="raised")
        self.value = tk.BooleanVar(self, False)
        self.bind("<Button-1>", self.toggle)
        self.draw_button()

    def draw_button(self):
        self.label = tk.Label(self, text='Off', bg='red', width=8)
        self.label.pack(pady=2, padx=2)

    def toggle(self, event):
        self.value.set(not self.value.get())
        self.label.config(text='On' if self.value.get() else 'Off', bg='green' if self.value.get() else 'red')


class MyApp:
    def __init__(self, root):
        self.root = root
        root.title("Pugh Matrix")
        self.root.geometry("700x500")
        self.create_menu()
        self.create_content_frame()

        # Create a style object
        style = ttk.Style(root)
        # Set the theme
        style.theme_use('clam')  # 'alt', 'clam', 'classic', 'default', 'vista'

        # Configure Label Style
        style.configure('TLabel', font=('Helvetica', 12), background='#f0f0f0', foreground='black')

        # Entry Widget Styling
        style.configure('TEntry', font=('Helvetica', 12), padding=5, relief='flat', background='white')


        # Configure Button Style
        style.configure('TButton', font=('Helvetica', 12), padding=1, relief='flat', background='#0066FF', foreground='white')
        style.map('TButton', 
            foreground=[('pressed', 'white'), ('active', 'white')],
            background=[('pressed', '!disabled', 'black'), ('active', '#0057d9')],
            relief=[('pressed', 'sunken'), ('!pressed', 'flat')])

        # Label styling
        style.configure('TLabel', font=('Helvetica', 12), background='lightgray')

        # Entry styling with padding
        style.configure('TEntry', padding=5)

        self.solution_data = [
            {'name': 'Solution 1', 'details': ''},
            {'name': 'Solution 2', 'details': ''}
        ]
        
        self.criteria_data = [
            {'name': 'Criteria 1', 'details': '', 'importance': 'Low'},
            {'name': 'Criteria 2', 'details': '', 'importance': 'Low'}
        ]

        self.solution_entries = []  # Entries for solution names
        self.criteria_entries = []  # Entries for criteria names
        self.importance_comboboxes = []  # Comboboxes for criteria importance

        self.show_input_solution_frame()

    def toggle_theme(self):
        current_theme = self.style.theme_use()
        new_theme = 'alt' if current_theme == 'clam' else 'clam'
        self.style.theme_use(new_theme)
        messagebox.showinfo("Theme Toggled", f"Switched to {'Dark' if new_theme == 'alt' else 'Light'} Theme")


    def create_content_frame(self):
        self.content_frame = tk.Frame(self.root, background='#f0f0f0', pady=5, padx=5)
        self.content_frame.pack(fill='both', expand=True, padx=10, pady=10)


    def collect_export_state_data(self):
        data = []
        headers = ['Criteria', 'Details', 'Importance']
        solution_headers = []
        for solution in self.solution_data:
            solution_headers.extend([solution['name'] + " State", solution['name'] + " Details"])

        headers.extend(solution_headers)

        for i, criterion in enumerate(self.criteria_data):
            row = [
                criterion['name'],
                criterion.get('details', ''),
                criterion['importance']
            ]
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i+1, column=j+1)[0]
                state = checkbox.get_state() if isinstance(checkbox, CycleCheckbutton) else " "
                solution_detail = solution.get('details', '')
                row.extend([state, solution_detail])
            data.append(row)
        return headers, data

    def import_state(self):
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Import from"
        )
        if filename:
            self.import_state_from_json(filename)
        if not filename:
            return  # User cancelled the dialog

    def import_state_from_json(self, filename):
        try:
            with open(filename, 'r') as file:
                state = json.load(file)
            
            # Clear the frame without updating entries from UI to model
            self.clear_content_frame(update_entries=False)

            self.solution_data = state.get('solutions', [])
            self.criteria_data = state.get('criteria', [])

            # Rebuild the UI components with the newly imported data
            self.show_input_solution_frame()
            self.show_input_criteria_frame()
            
            messagebox.showinfo("Import Success", "State imported successfully.")
        except Exception as e:
            messagebox.showerror("Import Failed", f"Failed to import data: {e}")


    def refresh_views(self):
        # Refreshes the UI for solution and criteria entries
        self.show_input_criteria_frame()
        self.show_input_solution_frame()

    def export_state(self):
        # Ensure all data entries are updated before exporting
        self.update_data_from_entries()

        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Export as"
        )
        if filename:
            self.export_state_to_json(filename)
        else:
            messagebox.showerror("Export Error", "Unsupported file format selected.")

    def export_results(self):
        scores = self.calculate_scores()
        data = [(name, score) for name, score in scores.items()]

        # Options for save file dialog to allow selecting file type
        file_types = [
            ("Excel files", "*.xlsx"),
            ("CSV files", "*.csv"),
            ("PDF files", "*.pdf"),
            ("All files", "*.*")
        ]
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",  # Default file extension
            filetypes=file_types,      # List of tuples specifying allowed file types
            title="Save results as"
        )

        if not filename:
            return  # User cancelled the dialog

        if filename.endswith('.xlsx'):
            self.export_to_xlsx(scores, filename)
        elif filename.endswith('.csv'):
            self.export_to_csv(data, filename)
        elif filename.endswith('.pdf'):
            self.export_to_pdf(data, filename)
        else:
            messagebox.showerror("Export Error", "Unsupported file format selected.")

    def export_state_to_json(self, filename='state.json'):
        state = {
            'criteria': self.criteria_data,
            'solutions': self.solution_data,
            # Add other relevant data here
        }
        try:
            with open(filename, 'w') as file:
                json.dump(state, file, indent=4)
            messagebox.showinfo("Export Success", f"State exported successfully to {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def export_to_csv(self, data, filename):
        try:
            with open(filename, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(['Solution', 'Score'])
                for name, score in data:
                    writer.writerow([name, score])
            messagebox.showinfo("Export Success", f"Results exported successfully to CSV at {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def export_to_xlsx(self, scores, filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["Solution", "Score"])
            sorted_scores = sorted(scores.items(), key=lambda item: (-item[1], item[0]))
            for solution_name, score in sorted_scores:
                ws.append([solution_name, score])
            wb.save(filename)
            messagebox.showinfo("Export Success", f"Results exported successfully to XLSX at {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def prepare_data_for_export(self):
        # Calculate scores
        scores = self.calculate_scores()
        # Convert all scores to integers if they aren't already
        prepared_data = [(name, int(score)) for name, score in scores.items()]
        return prepared_data

    def export_to_pdf(self, data, filename):
        try:
            pdf = PDF(title='Results')
            pdf.add_page()
            pdf.set_font('Arial', '', 12)
            pdf.cell(40, 10, 'Solution', 0, 0)  # Adjusted cell width
            pdf.cell(0, 10, 'Score', 0, 1, 'R')  # 'R' aligns right

            sorted_data = sorted(data, key=lambda x: (-x[1], x[0]))
            for name, score in sorted_data:
                pdf.cell(40, 10, name, 0, 0)
                pdf.cell(0, 10, str(score), 0, 1, 'R')

            pdf.output(filename)
            messagebox.showinfo("Export Success", f"Results exported successfully to PDF as {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def create_menu(self):
        menubar = tk.Menu(self.root, background='lightgrey', foreground='black', activebackground='#0057d9', activeforeground='white')
        fileMenu = tk.Menu(menubar, tearoff=0, background='lightgrey', foreground='black')
        fileMenu.add_command(label="Open New", command=self.prompt_save_before_reset)
        fileMenu.add_command(label="Export State", command=self.export_state)
        fileMenu.add_command(label="Import State", command=self.import_state)
        fileMenu.add_separator()
        fileMenu.add_command(label="Help", command=lambda: messagebox.showinfo("Help", "Not supported yet!"))
        fileMenu.add_separator()
        fileMenu.add_command(label="Close", command=self.root.quit)
        menubar.add_cascade(label="File", menu=fileMenu)

        viewMenu = tk.Menu(menubar, tearoff=0, background='lightgrey', foreground='black')
        viewMenu.add_command(label="Input Solution", command=self.show_input_solution_frame)
        viewMenu.add_command(label="Input Criteria", command=self.show_input_criteria_frame)
        viewMenu.add_command(label="View Calculation", command=self.show_input_calculation_frame)
        menubar.add_cascade(label="View", menu=viewMenu)
        self.root.config(menu=menubar)


    def prompt_save_before_reset(self):
        response = messagebox.askyesnocancel("Save Current State", "Do you want to save the current state before opening a new session?")
        if response:  # Yes, save the state
            self.export_state()
            self.reset_application()  # Reset application after saving
        elif response is False:  # No, do not save
            self.reset_application()

    def reset_application(self):
        # Clear existing data
        self.solution_data.clear()
        self.criteria_data.clear()
        self.clear_content_frame()  # Clears the content frame

        # Populate with initial default data
        self.solution_data = [
            {'name': 'Solution 1', 'details': ''},
            {'name': 'Solution 2', 'details': ''}
        ]

        self.criteria_data = [
            {'name': 'Criteria 1', 'details': '', 'importance': 'Low'},
            {'name': 'Criteria 2', 'details': '', 'importance': 'Low'}
        ]

        # Re-initialize GUI components with default data
        self.show_input_solution_frame()  # This method should re-render the solution input section

    def perform_csv_export(self, scores):
        # Prepare data for export
        data = [(name, score) for name, score in scores.items()]
        self.export_to_csv(data, filename="Results.csv")

    def perform_xlsx_export(self, scores):
        # Prepare data for export
        self.export_to_xlsx(scores, filename="Results.xlsx")

    def perform_pdf_export(self, scores):
        # Prepare data for export
        data = [(name, score) for name, score in scores.items()]
        self.export_to_pdf(data, filename="Results.pdf")


    def get_scores(self):
        # Logic to collect scores from your application's data
        return [(solution['name'], score) for solution, score in zip(self.solution_data, self.calculate_scores())]

    def calculate_scores(self):
        scores = {}
        importance_scores = {"Low": 1, "Medium": 2, "High": 3}
        
        # Initialize scores dictionary with solution names
        for solution in self.solution_data:
            scores[solution['name']] = 0
        
        for i, criterion in enumerate(self.criteria_data):
            importance = criterion['importance']
            importance_score = importance_scores[importance]
            
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i + 1, column=j + 1)[0]
                if isinstance(checkbox, CycleCheckbutton):
                    # Multiply the checkbox score by the importance score and add to the solution's total
                    scores[solution['name']] += checkbox.get_score() * importance_score
        
        return scores

    def display_results(self, scores):
        figure = plt.Figure(figsize=(6, 5), dpi=100)
        ax = figure.add_subplot(111)
        data = [score for _, score in scores.items()]
        labels = [name for name, _ in scores.items()]
        ax.bar(labels, data, color='blue')
        ax.set_title('Solution Scores')
        canvas = FigureCanvasTkAgg(figure, self.root)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack()
        canvas.draw()

    def calculate_and_display(self):
        scores = self.calculate_scores()
        self.display_results(scores)


    def clear_content_frame(self, update_entries=True):
        # Clears the content frame, optionally updating entries to the data model before clearing
        if update_entries:
            self.update_data_from_entries()  # Update data model from entries if required

        self.criteria_entries.clear()
        self.importance_comboboxes.clear()
        self.solution_entries.clear()
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        self.reset_grid_configuration()  # Reset configuration after clearing widgets
        self.content_frame.update_idletasks()  # Force update


    def reset_grid_configuration(self):
        # Clear any previous grid configuration thoroughly
        current_rows, current_cols = self.content_frame.grid_size()
        for i in range(current_rows):
            self.content_frame.grid_rowconfigure(i, weight=0, minsize=0)
        for j in range(current_cols):
            self.content_frame.grid_columnconfigure(j, weight=0, minsize=0)

        # Apply new configurations based on current data
        for i in range(len(self.criteria_data) + 1):  # +1 for headers
            self.content_frame.grid_rowconfigure(i, weight=1)
        for j in range(len(self.solution_data) + 1):
            self.content_frame.grid_columnconfigure(j, weight=1)

    def update_data_from_entries(self):
        for i, entry in enumerate(self.criteria_entries):
            if i < len(self.criteria_data):
                self.criteria_data[i]['name'] = entry.get()
                self.criteria_data[i]['importance'] = self.importance_comboboxes[i].get()
                details_entry = entry.master.children['details_entry']
                self.criteria_data[i]['details'] = details_entry.get()

        for i, entry in enumerate(self.solution_entries):
            if i < len(self.solution_data):
                self.solution_data[i]['name'] = entry.get()
                # Capture details from the adjacent details entry widget
                details_entry = entry.master.children['details_entry']
                self.solution_data[i]['details'] = details_entry.get()

    def show_input_solution_frame(self):
        # Builds the input area for solution entries dynamically based on current data
        self.clear_content_frame()
        solution_container = tk.Frame(self.content_frame)
        solution_container.pack(fill='x', expand=True, pady=10)

        header_frame = tk.Frame(solution_container)
        header_frame.pack(fill='x', expand=True)
        ttk.Label(header_frame, text="Solutions", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Details", style='DataLabel.TLabel').pack(side='left', expand=True)

        # Clear previous entries list to avoid duplications
        self.solution_entries = []
        for index, solution in enumerate(self.solution_data):
            self.add_solution_row(solution_container, solution, index)

        action_frame = tk.Frame(self.content_frame)
        action_frame.pack(fill='x', pady=10)
        ttk.Button(action_frame, text="Add Solution", command=self.add_solution).pack(side='left', padx=10)
        ttk.Button(action_frame, text="Delete Solution", command=self.delete_solution).pack(side='left')

        self.reset_grid_configuration()

    def add_solution_row(self, container, solution, index):
        # Adds a single solution row in the UI
        row_frame = tk.Frame(container)
        row_frame.pack(fill='x', expand=True, padx=5, pady=5)
        entry = tk.Entry(row_frame)
        entry.insert(0, solution['name'])
        entry.pack(side='left', fill='x', expand=True)
        details_entry = tk.Entry(row_frame, name='details_entry')
        details_entry.insert(0, solution.get('details', ''))
        details_entry.pack(side='left', fill='x', expand=True)
        self.solution_entries.append(entry)  # Keep track of the entry for later use

    def add_solution(self):
        if len(self.solution_data) < 13:
            self.solution_data.append({'name': '', 'details': ''})
            self.show_input_solution_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Limit Reached", "A maximum of 13 solution rows are allowed.")

    def delete_solution(self):
        if len(self.solution_data) > 1:
            self.solution_data.pop()
            self.show_input_solution_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Minimum Requirement", "At least one solution must be present.")

    def show_input_criteria_frame(self):
        # Builds the input area for criteria entries dynamically based on current data
        self.clear_content_frame()
        criteria_container = tk.Frame(self.content_frame)
        criteria_container.pack(fill='x', expand=True, pady=10)

        header_frame = tk.Frame(criteria_container)
        header_frame.pack(fill='x', expand=True)
        ttk.Label(header_frame, text="Criteria", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Details", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Importance", style='DataLabel.TLabel').pack(side='left', expand=True)

        # Clear previous entries list to avoid duplications
        self.criteria_entries = []
        self.importance_comboboxes = []
        for index, criterion in enumerate(self.criteria_data):
            self.add_criteria_row(criteria_container, criterion, index)

        action_frame = tk.Frame(self.content_frame)
        action_frame.pack(fill='x', pady=10)
        ttk.Button(action_frame, text="Add Criteria", command=self.add_criteria).pack(side='left', padx=10)
        ttk.Button(action_frame, text="Delete Criteria", command=self.delete_criteria).pack(side='left')

        self.reset_grid_configuration()

    def add_criteria_row(self, container, criterion, index):
        # Adds a single criteria row in the UI
        row_frame = tk.Frame(container)
        row_frame.pack(fill='x', expand=True, padx=5, pady=5)
        
        entry = tk.Entry(row_frame)
        entry.insert(0, criterion['name'])
        entry.pack(side='left', fill='x', expand=True)
        
        details_entry = tk.Entry(row_frame, name='details_entry')
        details_entry.insert(0, criterion.get('details', ''))
        details_entry.pack(side='left', fill='x', expand=True)
        
        # Importance combobox setup
        importance_combobox = ttk.Combobox(row_frame, values=["Low", "Medium", "High"], state="readonly")
        importance_combobox.set(criterion['importance'])
        importance_combobox.pack(side='left', expand=True)
        
        self.criteria_entries.append(entry)  # Keep track of the entry for later use
        self.importance_comboboxes.append(importance_combobox)  # Keep track of the combobox for later use

    def add_criteria(self):
        if len(self.criteria_data) < 12:
            self.criteria_data.append({'name': '', 'importance': 'Low'})
            self.show_input_criteria_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Limit Reached", "A maximum of 12 criteria rows are allowed.")

    def delete_criteria(self):
        if len(self.criteria_data) > 1:
            self.criteria_data.pop()
            self.show_input_criteria_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Minimum Requirement", "At least one criteria must be present.")

    def show_input_calculation_frame(self):
        self.clear_content_frame()  # Clear previous widgets
        if not self.criteria_data or not self.solution_data:
            messagebox.showinfo("Info", "No data to display in calculation view.")
            return

        # Re-create the grid for the current state
        for i, criterion in enumerate(self.criteria_data):
            tk.Label(self.content_frame, text=criterion['name'], borderwidth=1, relief="solid", wraplength=100).grid(row=i+1, column=0, sticky="nsew")

        for j, solution in enumerate(self.solution_data):
            tk.Label(self.content_frame, text=solution['name'], borderwidth=1, relief="solid", wraplength=100).grid(row=0, column=j+1, sticky="nsew")
            
            for i, criterion in enumerate(self.criteria_data):
                state = criterion.get('states', {}).get(solution['name'], " ")
                checkbutton = CycleCheckbutton(
                    self.content_frame,
                    state_change_callback=lambda state, c=criterion, s=solution['name']: self.update_checkbox_state(c, s, state)
                )
                checkbutton.set_state(state)
                checkbutton.grid(row=i+1, column=j+1, sticky="nsew")

        # Apply new grid configurations
        for i in range(len(self.criteria_data) + 1):
            self.content_frame.grid_rowconfigure(i, weight=1)
            
        for j in range(len(self.solution_data) + 1):
            self.content_frame.grid_columnconfigure(j, weight=1)

        # Add a calculate button if it's part of your design
        calculate_button = ttk.Button(self.content_frame, text="Calculate", command=self.calculate)
        calculate_button.grid(row=len(self.criteria_data) + 1, columnspan=len(self.solution_data) + 1, sticky="ew")

        self.reset_grid_configuration()

    def update_checkbox_state(self, criterion, solution_name, state):
        # This method updates the state in the model
        criterion.setdefault('states', {})[solution_name] = state

    def calculate(self):
        scores = {}
        importance_scores = {"Low": 1, "Medium": 2, "High": 3}
        
        # Initialize scores dictionary with solution names
        for solution in self.solution_data:
            scores[solution['name']] = 0
        
        for i, criterion in enumerate(self.criteria_data):
            importance = criterion['importance']
            importance_score = importance_scores[importance]
            
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i + 1, column=j + 1)[0]
                if isinstance(checkbox, CycleCheckbutton):
                    # Multiply the checkbox score by the importance score and add to the solution's total
                    added_score = int(checkbox.get_score()) if isinstance(checkbox.get_score(), str) else checkbox.get_score()
                    scores[solution['name']] += added_score * importance_score

        # Create a new Toplevel window to display results
        result_window = tk.Toplevel(self.root)
        result_window.title("Results")
        result_window.geometry("400x400")
        
        # Sort solutions by score descending, and by name alphabetically if scores are the same
        sorted_solutions = sorted(scores.items(), key=lambda item: (-item[1], item[0]))
        
        # Display each solution and its score
        for i, (solution_name, score) in enumerate(sorted_solutions):
            result_label = tk.Label(result_window, text=f"{solution_name}: {score}", font=("Arial", 14))
            result_label.pack(pady=(20, 0) if i == 0 else (10, 0))

        # Buttons frame
        buttons_frame = tk.Frame(result_window)
        buttons_frame.pack(pady=20)

        # Button for exporting to CSV
        export_button = ttk.Button(buttons_frame, text="Export Results", command=self.export_results)
        export_button.pack(side='left', padx=10)

if __name__ == "__main__":
    root = tk.Tk()
    app = MyApp(root)
    root.mainloop()