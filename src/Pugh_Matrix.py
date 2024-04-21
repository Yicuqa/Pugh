import json
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from fpdf import FPDF
import os
import sys

#https://stackoverflow.com/questions/31836104/pyinstaller-and-onefile-how-to-include-an-image-in-the-exe-file
def resource_path(relative_path):
    # Get absolute path to resource, works for dev and for PyInstaller
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

Logo = resource_path("src\\dist\\assets\\pugh_icon.ico")

class PDF(FPDF):
    def __init__(self, title=None):
        super().__init__()
        self.title = title

    def header(self):
        if self.title:
            self.set_font('Arial', 'B', 12)
            self.cell(0, 10, self.title, 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

class CycleCheckbutton(tk.Label):
    def __init__(self, master=None, state_change_callback=None, **kw):
        super().__init__(master, **kw)
        self.configure(font=('Helvetica', 10), wraplength=100, bg='#f0f0f0', fg='black', relief='flat', pady=1, padx=1)
        self.values = [" ", "+", "-", "S"]
        self.current_value = 0
        self.configure(text=self.values[self.current_value], borderwidth=1, relief="groove")
        self.bind("<Button-1>", self.cycle_value)
        self.state_change_callback = state_change_callback

    def cycle_value(self, event=None):
        self.current_value = (self.current_value + 1) % len(self.values)
        self.configure(text=self.values[self.current_value])
        # Invoke the callback if it's set
        if self.state_change_callback:
            self.state_change_callback(self.get_state())

    def get_state(self):
        return self.values[self.current_value]

    def set_state(self, state):
        if state in self.values:
            self.current_value = self.values.index(state)
            self.configure(text=state)

    def get_score(self):
        value_scores = {"+": 1, "-": -1, "S": 0, " ": 0}
        return value_scores[self.values[self.current_value]]

class MyApp:
    def __init__(self, root):
        self.root = root
        root.title("Pugh Matrix")
        self.root.geometry("700x500")
        self.create_menu()
        self.create_content_frame()

        Logo_path = resource_path("assets\\small_icon.ico")
        root.iconbitmap(r"{}".format(Logo_path))  # Use raw string notation

        # Create a style object
        style = ttk.Style(root)
        
        # Set the theme
        style.theme_use('clam')  # 'alt', 'clam', 'classic', 'default', 'vista'

        # Configure Label Style
        style.configure('TLabel', font=('Helvetica', 12), background='#f0f0f0', foreground='black')

        # Entry Widget Styling
        style.configure('TEntry', font=('Helvetica', 12), padding=5, relief='flat', background='white')

        # Configure Button Style
        style.configure('TButton', font=('Helvetica', 12), padding=1, relief='flat', background='#0066FF', foreground='white')
        style.map('TButton', 
            foreground=[('pressed', 'white'), ('active', 'white')],
            background=[('pressed', '!disabled', 'black'), ('active', '#0057d9')],
            relief=[('pressed', 'sunken'), ('!pressed', 'flat')])

        # Entry styling with padding
        style.configure('TEntry', padding=5)

        self.solution_data = [
            {'name': 'Baseline', 'details': 'If compared with existing: S, + or -'},
            {'name': 'Solution 1', 'details': ''}
        ]
        
        self.criteria_data = [
            {'name': 'Criteria 1', 'details': '', 'importance': 'Low'},
            {'name': 'Criteria 2', 'details': '', 'importance': 'Low'}
        ]

        self.solution_entries = []  # Entries for solution names
        self.criteria_entries = []  # Entries for criteria names
        self.importance_comboboxes = []  # Comboboxes for criteria importance

        self.show_input_solution_frame()

    def create_content_frame(self):
        self.content_frame = tk.Frame(self.root, background='#f0f0f0', pady=5, padx=5)
        self.content_frame.pack(fill='both', expand=True, padx=10, pady=10)

    def collect_export_state_data(self):
        data = []
        headers = ['Criteria', 'Details', 'Importance']
        solution_headers = []
        for solution in self.solution_data:
            solution_headers.extend([solution['name'] + " State", solution['name'] + " Details"])

        headers.extend(solution_headers)

        for i, criterion in enumerate(self.criteria_data):
            row = [
                criterion['name'],
                criterion.get('details', ''),
                criterion['importance']
            ]
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i+1, column=j+1)[0]
                state = checkbox.get_state() if isinstance(checkbox, CycleCheckbutton) else " "
                solution_detail = solution.get('details', '')
                row.extend([state, solution_detail])
            data.append(row)
        return headers, data

    def import_state(self):
        filename = filedialog.askopenfilename(
            filetypes=[("JSON", "*.json"), ("All files", "*.*")],
            title="Import from"
        )
        if filename:
            self.import_state_from_json(filename)
        if not filename:
            return  # User cancelled the dialog

    def import_state_from_json(self, filename):
        with open(filename, 'r') as file:
            state = json.load(file)
        
        # Clear the frame without updating entries from UI to model
        self.clear_content_frame(update_entries=False)

        self.solution_data = state.get('solutions', [])
        self.criteria_data = state.get('criteria', [])

        # Rebuild the UI components with the newly imported data
        self.show_input_criteria_frame()
        self.show_input_solution_frame()

    def refresh_views(self):
        # Refreshes the UI for solution and criteria entries
        self.show_input_criteria_frame()
        self.show_input_solution_frame()

    def export_state(self):
        # Ensure all data entries are updated before exporting
        self.update_data_from_entries()

        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON", "*.json"), ("All files", "*.*")],
            title="Export state"
        )

        if not filename:
            return  # User cancelled the dialog
        
        if filename:
            self.export_state_to_json(filename)
        else:
            messagebox.showerror("Export Error", "Unsupported file format selected.")

    def export_results(self):
        scores = self.calculate_scores()
        data = [(name, score) for name, score in scores.items()]

        # Options for save file dialog to allow selecting file type
        file_types = [
            ("Excel Workbook", "*.xlsx"),
            ("CSV (Comma delimited)", "*.csv"),
            ("PDF", "*.pdf"),
            ("All files", "*.*")
        ]
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",  # Default file extension
            filetypes=file_types,      # List of tuples specifying allowed file types
            title="Export results"
        )

        if not filename:
            return  # User cancelled the dialog

        if filename.endswith('.xlsx'):
            self.export_to_xlsx(scores, filename)
        elif filename.endswith('.csv'):
            self.export_to_csv(data, filename)
        elif filename.endswith('.pdf'):
            self.export_to_pdf(data, filename)
        else:
            messagebox.showerror("Export Error", "Unsupported file format selected.")

    def export_state_to_json(self, filename='state.json'):
        state = {
            'criteria': self.criteria_data,
            'solutions': self.solution_data,
            # Add other relevant data here
        }
        try:
            with open(filename, 'w') as file:
                json.dump(state, file, indent=4)
            messagebox.showinfo("Export Success", f"State exported successfully to {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def export_to_csv(self, data, filename):
        try:
            with open(filename, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerow(['Solution', 'Score'])
                for name, score in data:
                    writer.writerow([name, score])
            messagebox.showinfo("Export Success", f"Results exported successfully to CSV at {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def export_to_xlsx(self, scores, filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["Solution", "Score"])
            sorted_scores = sorted(scores.items(), key=lambda item: (-item[1], item[0]))
            for solution_name, score in sorted_scores:
                ws.append([solution_name, score])
            wb.save(filename)
            messagebox.showinfo("Export Success", f"Results exported successfully to XLSX at {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def prepare_data_for_export(self):
        # Calculate scores
        scores = self.calculate_scores()
        # Convert all scores to integers if they aren't already
        prepared_data = [(name, int(score)) for name, score in scores.items()]
        return prepared_data

    def export_to_pdf(self, data, filename):
        try:
            pdf = PDF(title='Results')
            pdf.add_page()
            pdf.set_font('Arial', '', 12)
            pdf.cell(40, 10, 'Solution', 0, 0)  # Adjusted cell width
            pdf.cell(0, 10, 'Score', 0, 1, 'R')  # 'R' aligns right

            sorted_data = sorted(data, key=lambda x: (-x[1], x[0]))
            for name, score in sorted_data:
                pdf.cell(40, 10, name, 0, 0)
                pdf.cell(0, 10, str(score), 0, 1, 'R')

            pdf.output(filename)
            messagebox.showinfo("Export Success", f"Results exported successfully to PDF as {filename}.")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data: {e}")

    def create_menu(self):
        menubar = tk.Menu(self.root, background='lightgrey', foreground='black', activebackground='#0057d9', activeforeground='white')
        fileMenu = tk.Menu(menubar, tearoff=0, background='#f0f0f0', foreground='black')
        fileMenu.add_command(label="Open New", command=self.prompt_save_before_reset)
        fileMenu.add_command(label="Export State", command=self.export_state)
        fileMenu.add_command(label="Import State", command=self.import_state)
        fileMenu.add_separator()
        fileMenu.add_command(label="Close", command=self.root.quit)
        menubar.add_cascade(label="File", menu=fileMenu)

        viewMenu = tk.Menu(menubar, tearoff=0, background='#f0f0f0', foreground='black')
        viewMenu.add_command(label="Solution", command=self.show_input_solution_frame)
        viewMenu.add_command(label="Criteria", command=self.show_input_criteria_frame)
        viewMenu.add_command(label="Calculation", command=self.show_input_calculation_frame)
        menubar.add_cascade(label="View", menu=viewMenu)

        menubar.add_command(label="Instructions", command=self.show_instructions_frame)
        self.root.config(menu=menubar)


    def prompt_save_before_reset(self):
        response = messagebox.askyesnocancel("Save Current State", "Do you want to save the current state before opening a new session?")
        if response:  # Yes, save the state
            self.export_state()
            self.reset_application()  # Reset application after saving
        elif response is False:  # No, do not save
            self.reset_application()

    def reset_application(self):
        # Clear existing data
        self.solution_data.clear()
        self.criteria_data.clear()
        self.clear_content_frame(update_entries=False)  # Clears the content frame
        
        # Reset and reconfigure grid
        self.reset_grid_configuration()
        self.content_frame.update_idletasks()  # Process all pending geometry tasks

        # Populate with initial default data
        self.solution_data = [
            {'name': 'Baseline', 'details': 'If compared with existing: S, + or -'},
            {'name': 'Solution 1', 'details': ''}
        ]
        self.criteria_data = [
            {'name': 'Criteria 1', 'details': '', 'importance': 'Low'},
            {'name': 'Criteria 2', 'details': '', 'importance': 'Low'}
        ]

        # Re-initialize GUI components with default data
        self.show_input_criteria_frame()
        self.show_input_solution_frame()


    def perform_csv_export(self, scores):
        # Prepare data for export
        data = [(name, score) for name, score in scores.items()]
        self.export_to_csv(data, filename="Results.csv")

    def perform_xlsx_export(self, scores):
        # Prepare data for export
        self.export_to_xlsx(scores, filename="Results.xlsx")

    def perform_pdf_export(self, scores):
        # Prepare data for export
        data = [(name, score) for name, score in scores.items()]
        self.export_to_pdf(data, filename="Results.pdf")


    def get_scores(self):
        # Logic to collect scores from your application's data
        return [(solution['name'], score) for solution, score in zip(self.solution_data, self.calculate_scores())]

    def calculate_scores(self):
        scores = {}
        importance_scores = {"Low": 1, "Medium": 2, "High": 3}
        
        # Initialize scores dictionary with solution names
        for solution in self.solution_data:
            scores[solution['name']] = 0
        
        for i, criterion in enumerate(self.criteria_data):
            importance = criterion['importance']
            importance_score = importance_scores[importance]
            
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i + 1, column=j + 1)[0]
                if isinstance(checkbox, CycleCheckbutton):
                    # Multiply the checkbox score by the importance score and add to the solution's total
                    scores[solution['name']] += checkbox.get_score() * importance_score
        
        return scores

    def clear_content_frame(self, update_entries=True):
        # Clears the content frame, optionally updating entries to the data model before clearing
        if update_entries:
            self.update_data_from_entries()  # Update data model from entries if required

        self.criteria_entries.clear()
        self.importance_comboboxes.clear()
        self.solution_entries.clear()
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        self.reset_grid_configuration()  # Reset configuration after clearing widgets
        self.content_frame.update_idletasks()  # Force update

    def reset_grid_configuration(self):
        # First, reset all possible grid configurations to default
        max_rows = 15
        max_cols = 15
        for i in range(max_rows):
            self.content_frame.grid_rowconfigure(i, weight=0, minsize=0)
        for j in range(max_cols):
            self.content_frame.grid_columnconfigure(j, weight=0, minsize=0)

        # Now, configure only the necessary rows and columns
        for i in range(len(self.criteria_data) + 1):  # +1 for headers
            self.content_frame.grid_rowconfigure(i, weight=1)
        for j in range(len(self.solution_data) + 1):
            self.content_frame.grid_columnconfigure(j, weight=1)


    def update_data_from_entries(self):
        for i, entry in enumerate(self.criteria_entries):
            if i < len(self.criteria_data):
                self.criteria_data[i]['name'] = entry.get()
                self.criteria_data[i]['importance'] = self.importance_comboboxes[i].get()
                details_entry = entry.master.children['details_entry']
                self.criteria_data[i]['details'] = details_entry.get()

        for i, entry in enumerate(self.solution_entries):
            if i < len(self.solution_data):
                self.solution_data[i]['name'] = entry.get()
                # Capture details from the adjacent details entry widget
                details_entry = entry.master.children['details_entry']
                self.solution_data[i]['details'] = details_entry.get()


    def show_instructions_frame(self):
        # Create a new Toplevel window to display instructions
        instructions_window = tk.Toplevel(self.root)
        instructions_window.title("Instructions")
        instructions_window.geometry("620x400")

        Logo_path = resource_path("assets\\small_icon.ico")
        instructions_window.iconbitmap(r"{}".format(Logo_path))  # Use raw string notation

        # Create a frame for the scrollbar and text widget
        text_frame = ttk.Frame(instructions_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Create a scrollbar
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Create a text widget with left alignment
        text_widget = tk.Text(text_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, font=('Helvetica', 10), padx=10, pady=10, borderwidth=2, relief="groove")
        text_widget.pack(fill=tk.BOTH, expand=True)

        # Configure scrollbar
        scrollbar.config(command=text_widget.yview)

        # Define tags for different styling
        text_widget.tag_configure('title', justify='center', font=('Helvetica', 12, 'bold'), spacing3=0)
        text_widget.tag_configure('no_indent', justify='left', lmargin1=0, rmargin=0, spacing1=0, spacing3=0)

        # Title insertion
        text_widget.insert(tk.END, 'Introduction to Pugh Matrix\n\n', 'title')

        # Insert the help text before the instructions subtitle
        pre_instructions_text = """
        The Pugh Matrix is an effective tool for comparing a range of options — whether it be components, situations, solutions, vehicles or even interview candidates — against established criteria. This method facilitates a shift in decision-making from subjective impressions to an objective, data-focused strategy. Utilizing the Pugh Matrix helps reduce unconscious bias, ensuring that decisions are grounded in explicitly defined and evaluated criteria.

        Before we begin with the instructions:
        In the menu under ‘File’, you can export and import the current state of filled in solutions, criteria and even the checkboxes in the calculation view. If you’re in a hurry and don’t have time to finish the form, simply export the state and import it at a later time.

        You’re also able to export the result as either a PDF, CSV or XLSX file.
        
        
        """

        # Make sure to strip leading whitespace from strings
        formatted_text = "\n".join(line.strip() for line in pre_instructions_text.split("\n"))
        # Insert pre-instructions text with 'no_indent' tag
        text_widget.insert(tk.END, formatted_text, 'no_indent')

        # Insert the subtitle "Instructions"
        text_widget.insert(tk.END, '\nInstructions\n', 'title')

        # Insert the help text after the instructions subtitle
        instructions_text = """
        Step 1: Enter ‘Solution View’ and list possible solutions
        Objective: Compile a comprehensive list of components or solutions you intend to compare.
        Action: In the 'Solution View', add a brief name for each solution and include a detailed description to ensure all stakeholders have a clear understanding. This list will be automatically integrated into the ‘Calculation View’ for comparison.

        Step 2: Enter ‘Criteria View’ and define and detail a list of criteria
        Objective: Identify and document the criteria that are crucial for comparing the different solutions.
        Action: Update the 'Criteria View' with essential criteria and provide a detailed description of each to align understanding among all involved parties. These criteria will be automatically transferred to the corresponding section in the ‘Calculation View’.

        Step 3: Rate the Importance of each Criteria
        Objective: Assess and assign a significance level to each criteria to reflect its importance in the decision-making process.
        Action: Review the list of criteria in the ‘Criteria View’ and assign a rating from ‘Low’ to ‘High’, where ‘Low’ indicates lesser importance, ‘Medium’ indicates quite important, and ‘High’ signifies it's of critical importance.

        Step 4: Establish a Baseline for Comparison
        Objective: Set a baseline component for a straightforward comparison with other components.
        Action: If a baseline exists, fill in the satisfaction level using 'S' for Satisfied, '+' for Better Than Satisfied, and '-' for Not Satisfied, according to the agreed criteria. If there is no Baseline, one of the solution can be used as one instead.

        Step 5: Evaluate Each Component
        Objective: Apply the same evaluation criteria to each component under consideration.
        Action: Mark each component with 'S', '+', or '-', where 'S' stands for "Same as Baseline", '+' for "Better" and '-' for "Worse".

        Step 6: Interpret the Results
        Objective: Utilize the insights gained from the matrix to make informed decisions.
        Action: Press the ‘Calculate’ button in the ‘Calculation View’ in order to see the results. The solution with the highest points will be presented at the top and the lowest will be at the bottom. The more points a solution receive, the better the indication that the solution is of a more superior outcome.
        """
        # Make sure to strip leading whitespace from strings
        formatted_text2 = "\n".join(line.strip() for line in instructions_text.split("\n"))
        # Insert instruction text with 'no_indent' tag
        text_widget.insert(tk.END, formatted_text2, 'no_indent')
        text_widget.config(state=tk.DISABLED)  # Disable editing of the text widget


    def show_input_solution_frame(self):
        # Builds the input area for solution entries dynamically based on current data
        self.clear_content_frame()
        solution_container = tk.Frame(self.content_frame)
        solution_container.pack(fill='x', expand=True, pady=10)

        header_frame = tk.Frame(solution_container)
        header_frame.pack(fill='x', expand=True)
        ttk.Label(header_frame, text="Solutions", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Details", style='DataLabel.TLabel').pack(side='left', expand=True)

        # Clear previous entries list to avoid duplications
        self.solution_entries = []
        for index, solution in enumerate(self.solution_data):
            self.add_solution_row(solution_container, solution, index)

        action_frame = tk.Frame(self.content_frame)
        action_frame.pack(fill=tk.X, pady=10)

        # Create an inner frame to hold the buttons
        button_frame = tk.Frame(action_frame)
        button_frame.pack(pady=10, padx=10)

        # Add buttons to the button frame
        add_button = ttk.Button(button_frame, text="Add Solution", command=self.add_solution)
        add_button.pack(side=tk.LEFT, padx=10)

        delete_button = ttk.Button(button_frame, text="Delete Solution", command=self.delete_solution)
        delete_button.pack(side=tk.LEFT)

        # Center the button_frame within action_frame
        button_frame.pack_configure(side=tk.TOP, expand=True)

        self.reset_grid_configuration()

    def add_solution_row(self, container, solution, index):
        # Adds a single solution row in the UI
        row_frame = tk.Frame(container)
        row_frame.pack(fill='x', expand=True, padx=5, pady=5)
        entry = tk.Entry(row_frame)
        entry.insert(0, solution['name'])
        entry.pack(side='left', fill='x', expand=True)
        details_entry = tk.Entry(row_frame, name='details_entry')
        details_entry.insert(0, solution.get('details', ''))
        details_entry.pack(side='left', fill='x', expand=True)
        self.solution_entries.append(entry)  # Keep track of the entry for later use

    def add_solution(self):
        if len(self.solution_data) < 12:
            self.solution_data.append({'name': '', 'details': ''})
            self.show_input_solution_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Limit Reached", "A maximum of 12 solution rows are allowed.")

    def delete_solution(self):
        if len(self.solution_data) > 1:
            self.solution_data.pop()
            self.show_input_solution_frame()
            # Reset configuration for the now non-existent column
            self.reset_grid_configuration()
            # Additional resetting specific column (if needed)
            self.content_frame.grid_columnconfigure(len(self.solution_data)+1, minsize=0, weight=0)
        else:
            messagebox.showinfo("Minimum Requirement", "At least one solution must be present.")


    def show_input_criteria_frame(self):
        # Builds the input area for criteria entries dynamically based on current data
        self.clear_content_frame()
        criteria_container = tk.Frame(self.content_frame)
        criteria_container.pack(fill='x', expand=True, pady=10)

        header_frame = tk.Frame(criteria_container)
        header_frame.pack(fill='x', expand=True)
        ttk.Label(header_frame, text="Criteria", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Details", style='DataLabel.TLabel').pack(side='left', expand=True)
        ttk.Label(header_frame, text="Importance", style='DataLabel.TLabel').pack(side='left', expand=True)

        # Clear previous entries list to avoid duplications
        self.criteria_entries = []
        self.importance_comboboxes = []
        for index, criterion in enumerate(self.criteria_data):
            self.add_criteria_row(criteria_container, criterion, index)

        action_frame = tk.Frame(self.content_frame)
        action_frame.pack(fill=tk.X, pady=10)

        # Create an inner frame to hold the buttons
        button_frame = tk.Frame(action_frame)
        button_frame.pack(pady=10, padx=10)

        # Add buttons to the button frame
        add_button = ttk.Button(button_frame, text="Add Criteria", command=self.add_criteria)
        add_button.pack(side=tk.LEFT, padx=10)

        delete_button = ttk.Button(button_frame, text="Delete Criteria", command=self.delete_criteria)
        delete_button.pack(side=tk.LEFT)

        self.reset_grid_configuration()

    def add_criteria_row(self, container, criterion, index):
        # Adds a single criteria row in the UI
        row_frame = tk.Frame(container)
        row_frame.pack(fill='x', expand=True, padx=5, pady=5)
        
        entry = tk.Entry(row_frame)
        entry.insert(0, criterion['name'])
        entry.pack(side='left', fill='x', expand=True)
        
        details_entry = tk.Entry(row_frame, name='details_entry')
        details_entry.insert(0, criterion.get('details', ''))
        details_entry.pack(side='left', fill='x', expand=True)
        
        # Importance combobox setup
        importance_combobox = ttk.Combobox(row_frame, values=["Low", "Medium", "High"], state="readonly")
        importance_combobox.set(criterion['importance'])
        importance_combobox.pack(side='left', expand=True)
        
        self.criteria_entries.append(entry)  # Keep track of the entry for later use
        self.importance_comboboxes.append(importance_combobox)  # Keep track of the combobox for later use

    def add_criteria(self):
        if len(self.criteria_data) < 11:
            self.criteria_data.append({'name': '', 'importance': 'Low'})
            self.show_input_criteria_frame()
            self.reset_grid_configuration()
        else:
            messagebox.showinfo("Limit Reached", "A maximum of 11 criteria rows are allowed.")

    def delete_criteria(self):
        if len(self.criteria_data) > 1:
            self.criteria_data.pop()
            self.show_input_criteria_frame()
            # Reset configuration for the now non-existent row
            self.reset_grid_configuration()
            # Additional resetting specific row (if needed)
            self.content_frame.grid_rowconfigure(len(self.criteria_data)+1, minsize=0, weight=0)
        else:
            messagebox.showinfo("Minimum Requirement", "At least one criteria must be present.")


    def show_input_calculation_frame(self):
        self.clear_content_frame()  # Clear previous widgets
        if not self.criteria_data or not self.solution_data:
            messagebox.showinfo("Info", "No data to display in calculation view.")
            return

        # Re-create the grid for the current state
        for i, criterion in enumerate(self.criteria_data):
            tk.Label(self.content_frame, text=criterion['name'], borderwidth=1, relief="solid", wraplength=100).grid(row=i+1, column=0, sticky="nsew")

        for j, solution in enumerate(self.solution_data):
            tk.Label(self.content_frame, text=solution['name'], borderwidth=1, relief="solid", wraplength=100).grid(row=0, column=j+1, sticky="nsew")
            
            for i, criterion in enumerate(self.criteria_data):
                state = criterion.get('states', {}).get(solution['name'], " ")
                checkbutton = CycleCheckbutton(
                    self.content_frame,
                    state_change_callback=lambda state, c=criterion, s=solution['name']: self.update_checkbox_state(c, s, state)
                )
                checkbutton.set_state(state)
                checkbutton.grid(row=i+1, column=j+1, sticky="nsew")

        # Apply new grid configurations
        for i in range(len(self.criteria_data) + 1):
            self.content_frame.grid_rowconfigure(i, weight=1)
            
        for j in range(len(self.solution_data) + 1):
            self.content_frame.grid_columnconfigure(j, weight=1)

        # Add a calculate button if it's part of your design
        calculate_button = ttk.Button(self.content_frame, text="Calculate", command=self.calculate)
        calculate_button.grid(row=len(self.criteria_data) + 1, columnspan=len(self.solution_data) + 1, sticky="ew")

        self.reset_grid_configuration()

    def update_checkbox_state(self, criterion, solution_name, state):
        # This method updates the state in the model
        criterion.setdefault('states', {})[solution_name] = state

    def calculate(self):
        scores = {}
        importance_scores = {"Low": 1, "Medium": 2, "High": 3}
        
        # Initialize scores dictionary with solution names
        for solution in self.solution_data:
            scores[solution['name']] = 0
        
        for i, criterion in enumerate(self.criteria_data):
            importance = criterion['importance']
            importance_score = importance_scores[importance]
            
            for j, solution in enumerate(self.solution_data):
                checkbox = self.content_frame.grid_slaves(row=i + 1, column=j + 1)[0]
                if isinstance(checkbox, CycleCheckbutton):
                    # Multiply the checkbox score by the importance score and add to the solution's total
                    added_score = int(checkbox.get_score()) if isinstance(checkbox.get_score(), str) else checkbox.get_score()
                    scores[solution['name']] += added_score * importance_score

        # Create a new Toplevel window to display results
        result_window = tk.Toplevel(self.root)
        result_window.title("Results")
        result_window.geometry("400x400")
        
        Logo_path = resource_path("assets\\small_icon.ico")
        result_window.iconbitmap(r"{}".format(Logo_path))  # Use raw string notation
        
        # Sort solutions by score descending, and by name alphabetically if scores are the same
        sorted_solutions = sorted(scores.items(), key=lambda item: (-item[1], item[0]))
        
        # Display each solution and its score
        for i, (solution_name, score) in enumerate(sorted_solutions):
            result_label = tk.Label(result_window, text=f"{solution_name}: {score}", font=("Arial", 14))
            result_label.pack(pady=(20, 0) if i == 0 else (10, 0))

        # Buttons frame at the bottom
        buttons_frame = tk.Frame(result_window)
        buttons_frame.pack(side=tk.BOTTOM, pady=20)  # Now using side=tk.BOTTOM

        # Button for exporting to CSV
        export_button = ttk.Button(buttons_frame, text="Export Results", command=self.export_results)
        export_button.pack(side='left', padx=10)

if __name__ == "__main__":
    root = tk.Tk()
    app = MyApp(root)
    root.mainloop()